###############################################
# # Excel Translator
# # Version 1.0
# # 2024-06-19
###############################################
# # Description:
# # This Python script translates the text in an Excel file to a target language
# # using the Azure Cognitive Services Translator Text API.

import openpyxl
import os, requests, uuid
import datetime

# check for presence of environment variables
def check_env_var(var : str):
    if var not in os.environ:
        print(var + ' variable not set')
        # we allow you to set the variable at run time
        os.environ[var] = input('Enter ' + var + ': ')

def setup_language_api_access():
    # check for presence of environment variables
    check_env_var('TRANSLATOR_TEXT_SUBSCRIPTION_KEY')
    check_env_var('TRANSLATOR_TEXT_REGION')
    # using default endpoint of https://api.cognitive.microsofttranslator.com/
    # return object with values
    return {
        'subscription_key': os.environ['TRANSLATOR_TEXT_SUBSCRIPTION_KEY'],
        'region': os.environ['TRANSLATOR_TEXT_REGION'],
        'endpoint': 'https://api.cognitive.microsofttranslator.com/'
    }

# List the supported languages
def list_api_supported_languages(apiaccess):
    # If you encounter any issues with the base_url or path, make sure
    # that you are using the latest endpoint: https://docs.microsoft.com/azure/cognitive-services/translator/reference/v3-0-languages
    path = '/languages?api-version=3.0'
    constructed_url = apiaccess['endpoint'] + path

    headers = {
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }

    request = requests.get(constructed_url, headers=headers)
    response = request.json()
    # print each key in response json 
    print('Supported languages: ')
    for key in response['dictionary'].keys():
        print(key + ' : ' + response['dictionary'][key]['name'])


def translate_text(text : str, language : str, apiaccess) -> str:
    # If you encounter any issues with the base_url or path, make sure
    # that you are using the latest endpoint: https://docs.microsoft.com/azure/cognitive-services/translator/reference/v3-0-translate
    path = '/translate?api-version=3.0'
    params = '&from=en&to=' + language
    constructed_url = apiaccess['endpoint'] + path + params

    headers = {
        'Ocp-Apim-Subscription-Key': apiaccess['subscription_key'],
        'Ocp-Apim-Subscription-Region': apiaccess['region'],
        'Content-type': 'application/json',
        'X-ClientTraceId': str(uuid.uuid4())
    }

    body = [{
        'text' : text
    }]
    request = requests.post(constructed_url, headers=headers, json=body)
    response = request.json()

    return response[0]['translations'][0]['text']

def translate_excel_file(workbook, newworkbook, language, apiaccess):
    # Iterate through each worksheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        newsheet = newworkbook.create_sheet(title=sheet_name)
        
        # Iterate through each cell in the worksheet
        for row in sheet.iter_rows():
            for cell in row:
                # check if the cell is a string
                if not cell.data_type == 's':
                    continue
                # don't translate, just copy accross first header row
                if cell.row == 1:
                    newsheet[cell.coordinate] = cell.value
                    continue

                # Translate the text
                translated_text = translate_text(cell.value, language, apiaccess)
                # Write the translated text to the new worksheet
                newsheet[cell.coordinate] = translated_text
       
        # Close the Excel file
    workbook.close()

def main():

    # setup the language api access
    apiaccess = setup_language_api_access()
    # first list the supported languages
    list_api_supported_languages(apiaccess)

    language = input('Please Choose a target language using a 2 letter code from the list above: ')

    # choose a file from the local directory
    # print the current directory
    print('Current directory: ' + os.getcwd())
    # list the files in the current directory
    print('Files in current directory: ')
    count = 0
    # create a dictionary of files in the current directory
    dict_files = {}
    for file in os.listdir(os.getcwd()):
        if file.endswith('.xlsx'):
            dict_files[str(count)] = file
            print(str(count) + " : " + file)
            count += 1
        
    # choose a file
    file_to_translate = input('Please choose a file number from the list above: ')
    if not file_to_translate.isdigit():
        # generate an exception
        raise ValueError('Invalid file number')
    # Open the Excel file
    print('Translating ' + dict_files[file_to_translate])
    workbook = openpyxl.load_workbook(dict_files[file_to_translate])
    # Create a new Excel file
    newworkbook = openpyxl.Workbook()

    # Translate the Excel file and creates a new Excel Workbook object
    # with the translations
    translate_excel_file(workbook, newworkbook, language, apiaccess)
    
    # Save the new Excel workbook using the current date and time in the file name
    current_datetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    localPath = os.path.dirname(os.path.realpath(__file__))
    targetPath = os.path.join(localPath, 'aicards-translated-' + language + '-' + current_datetime + '.xlsx')
    newworkbook.save(targetPath)
    newworkbook.close()

if __name__ == "__main__":
    main()